"""Excel 导出/导入 — 一键备份/恢复。"""
import io
from sqlalchemy.orm import Session
from sqlalchemy import select, delete
from openpyxl import Workbook, load_workbook

from models.orm import (
    Tenant, Property, Contract, ContractItem, ContractImage,
    MoveInInspection, MoveInInspectionItem, MoveInInspectionImage,
    MeterReading, Bill, BillOtherFee, Payment, PaymentAllocation,
    DunningLog, MoveOutInspection, MoveOutInspectionItem, MoveOutInspectionImage, Settlement,
)

TABLES = [
    (Tenant, ["id_number", "name", "phone", "remark"]),
    (Property, ["id", "property_code", "name", "address", "status", "remark"]),
    (Contract, ["id", "contract_code", "property_id", "tenant_id_number",
     "residents_count", "monthly_rent", "deposit", "payment_method", "rent_due_day",
     "water_fee", "start_date", "end_date", "key_count", "status", "remark"]),
    (ContractItem, ["id", "contract_id", "item_name", "quantity", "sort_order"]),
    (ContractImage, ["id", "contract_id", "image_path", "sort_order"]),
    (MoveInInspection, ["id", "contract_id", "inspection_date",
     "meter_base_reading", "key_delivery_detail"]),
    (MoveInInspectionItem, ["id", "inspection_id",
     "item_name", "quantity", "status", "defect_remark"]),
    (MoveInInspectionImage, ["id", "inspection_id", "image_path", "sort_order"]),
    (MeterReading, ["id", "record_code", "contract_id", "reading_month",
     "last_reading", "current_reading", "electricity_amount", "reading_date", "remark"]),  # consumption is derived, skip
    (Bill, ["id", "bill_code", "contract_id", "bill_month", "rent", "water_fee",
     "electricity_fee", "total", "generated_date", "due_date", "status", "remark"]),
    (BillOtherFee, ["id", "bill_id", "fee_name", "amount", "remark"]),
    (Payment, ["id", "payment_date", "total_amount", "payment_method", "remark"]),
    (PaymentAllocation, ["id", "payment_id", "bill_id", "amount"]),
    (DunningLog, ["id", "bill_id", "dunning_time", "sms_content"]),
    (MoveOutInspection, ["id", "contract_id", "inspection_date",
     "meter_reading", "electricity_deduction", "key_return_status", "key_deduction", "remark"]),
    (MoveOutInspectionItem, ["id", "inspection_id",
     "item_name", "quantity", "status", "deduction_amount"]),
    (MoveOutInspectionImage, ["id", "inspection_id", "image_path", "sort_order"]),
    (Settlement, ["id", "contract_id", "deposit_total",
     "electricity_deduction", "item_damage_deduction", "item_missing_deduction",
     "key_deduction", "unpaid_bills_note", "unpaid_bills_total", "other_deduction",
     "refund_date", "refund_method", "remark"]),  # actual_refund is derived, skip
]

# 逆序删表（先子表后主表，避免外键冲突）
DROP_ORDER = [
    PaymentAllocation, DunningLog, BillOtherFee, Payment,
    Bill, MeterReading, MoveInInspectionItem, MoveInInspectionImage, MoveInInspection,
    MoveOutInspectionItem, MoveOutInspectionImage, MoveOutInspection, Settlement,
    ContractItem, ContractImage, Contract, Property, Tenant,
]


def export_to_excel(db: Session) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for model, columns in TABLES:
        ws = wb.create_sheet(title=model.__tablename__)
        ws.append(columns)
        for row in db.scalars(select(model)).all():
            ws.append([getattr(row, c, None) for c in columns])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_from_excel(db: Session, file_content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    stats = {}

    for model in DROP_ORDER:
        db.execute(delete(model))
    db.commit()

    for model, columns in TABLES:
        sheet_name = model.__tablename__
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = rows[0]
        col_idx = []
        for c in columns:
            try:
                col_idx.append(headers.index(c))
            except ValueError:
                col_idx.append(-1)
        count = 0
        for row in rows[1:]:
            if all(row[i] is None for i in col_idx if i >= 0):
                continue
            values = {}
            for c, i in zip(columns, col_idx):
                if i < 0:
                    continue
                val = row[i]
                if val is None:
                    if model.__table__.c[c].nullable:
                        continue
                    from sqlalchemy import Integer as SAInteger
                    col = model.__table__.c[c]
                    if isinstance(col.type, SAInteger):
                        val = 0
                    else:
                        val = ""
                values[c] = val
            if values:
                # Handle generated columns for old DB schemas
                if model is MeterReading and 'consumption' not in values:
                    values['consumption'] = (values.get('current_reading', 0) or 0) - (values.get('last_reading', 0) or 0)
                if model is Settlement and 'actual_refund' not in values:
                    deduct = sum(values.get(k, 0) or 0 for k in ['electricity_deduction', 'item_damage_deduction', 'item_missing_deduction', 'key_deduction', 'unpaid_bills_total', 'other_deduction'])
                    values['actual_refund'] = (values.get('deposit_total', 0) or 0) - deduct
                db.add(model(**values))
                count += 1
        stats[sheet_name] = count

    db.commit()
    stats["_total"] = sum(stats.values())
    return stats
